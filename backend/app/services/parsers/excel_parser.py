"""
Excel 文件解析器

使用 openpyxl 进行 XLSX 解析
"""

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from app.services.parsers.base import (
    BaseParser,
    DocumentMetadata,
    PageContent,
    ParsedDocument,
)

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Excel 文件解析器"""

    supported_extensions: List[str] = [".xlsx", ".xlsm", ".xltx", ".xltm"]
    supported_mime_types: List[str] = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroEnabled.12",
    ]

    async def parse(self, file_path: str) -> ParsedDocument:
        """
        解析 Excel 文件

        Args:
            file_path: 文件路径

        Returns:
            解析后的文档对象
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel parsing. "
                "Install it with: pip install openpyxl"
            )

        path = Path(file_path)
        stat = path.stat()

        wb = load_workbook(file_path, read_only=True, data_only=True)

        try:
            return self._parse_workbook(wb, path, stat.st_size)
        finally:
            wb.close()

    async def parse_bytes(self, data: bytes, filename: str) -> ParsedDocument:
        """
        从字节数据解析 Excel 文件

        Args:
            data: 文件字节数据
            filename: 文件名

        Returns:
            解析后的文档对象
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel parsing. "
                "Install it with: pip install openpyxl"
            )

        path = Path(filename)

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

        try:
            return self._parse_workbook(wb, path, len(data))
        finally:
            wb.close()

    def _parse_workbook(
        self,
        wb,
        path: Path,
        file_size: int,
    ) -> ParsedDocument:
        """
        解析 Excel 工作簿

        Args:
            wb: openpyxl Workbook 对象
            path: 文件路径
            file_size: 文件大小

        Returns:
            解析后的文档对象
        """
        # 提取属性
        props = wb.properties

        title = props.title or path.stem if props else path.stem
        author = props.creator if props else None
        created_at = props.created if props else None
        modified_at = props.modified if props else None

        # 提取每个工作表的内容
        pages = []
        all_content = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]

            sheet_content = self._extract_sheet_content(sheet)
            all_content.append(f"## {sheet_name}\n\n{sheet_content}")

            pages.append(
                PageContent(
                    page_number=sheet_idx + 1,
                    content=f"## {sheet_name}\n\n{sheet_content}",
                    tables=[f"sheet_{sheet_idx + 1}"],
                )
            )

        full_content = "\n\n".join(all_content)
        word_count = self.count_words(full_content)
        language = self.detect_language(full_content)

        metadata = DocumentMetadata(
            title=title,
            author=author,
            created_at=created_at,
            modified_at=modified_at,
            page_count=len(wb.sheetnames),
            file_type=path.suffix.lower(),
            file_size=file_size,
            word_count=word_count,
            language=language,
            custom_fields={
                "sheet_names": wb.sheetnames,
                "sheet_count": len(wb.sheetnames),
            },
        )

        return ParsedDocument(
            content=full_content,
            metadata=metadata,
            pages=pages,
        )

    def _extract_sheet_content(self, sheet) -> str:
        """
        提取工作表内容为文本

        Args:
            sheet: openpyxl Worksheet 对象

        Returns:
            工作表文本内容
        """
        rows_text = []

        for row in sheet.iter_rows(values_only=True):
            # 过滤空单元格
            cells = [str(cell) if cell is not None else "" for cell in row]

            # 跳过完全空的行
            if any(cell.strip() for cell in cells):
                rows_text.append(" | ".join(cells))

        return "\n".join(rows_text)
